import requests
import time
import csv
from openpyxl import load_workbook
from openpyxl import Workbook

url='https://maps.googleapis.com/maps/api/place/textsearch/json?query=street+food+in+uk&key=xxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
response = requests.get(url)
result = response.json()

count_page=1
place_all=[]
data_for_write_excel=[]
data_firts=[]
data_two=[]

def write_csv(data):
    with open('gm_40k.csv', 'a', encoding='utf-8',newline='') as f:#
        writer = csv.writer(f)
        writer.writerow([data['website'],
        data['address'],
        data['city'],
        data['state'],
        data['zip_cod'],
        data['phone'],
        data['count_of_rating'],
        data['rating'],
        data['url']])


def get_page_GM(url):
    response = requests.get(url)
    response_json = response.json()
    return response_json

def get_evry_place(place_id):
    #ChIJ8cfYodMEdkgRYyvN_4vXU4U
    if place_id != None:
        url ='https://maps.googleapis.com/maps/api/place/details/json?place_id={}&language=en&key=AIzaSyBLQ3gMYryO8fl_ZBRTpcIiDyNvNT7Maec'.format(place_id)
        #https://maps.googleapis.com/maps/api/place/details/json?place_id=ChIJ8cfYodMEdkgRYyvN_4vXU4U&language=en&key=AIzaSyBLQ3gMYryO8fl_ZBRTpcIiDyNvNT7Maec
        response_json = get_page_GM(url)
        print(place_id)
        website=response_json['result'].get('website', 'N/A')
        #address=response_json['result'].get('formatted_address', 'N/A').split(',')[0:-3]
        if len(response_json['result'].get('formatted_address', 'N/A'))==4:
            address=response_json['result'].get('formatted_address', 'N/A').split(',')[0]
        else:
            address = response_json['result'].get('formatted_address', 'N/A').split(',')[0]+response_json['result'].get('formatted_address', 'N/A').split(',')[1]
        city = response_json['result'].get('formatted_address', 'N/A').split(',')[-3]
        state = response_json['result'].get('formatted_address', 'N/A').split(',')[-2].strip().split(' ')[0]
        zip_cod = response_json['result'].get('formatted_address', 'N/A').split(',')[-2].strip().split(' ')[1]
        phone=response_json['result'].get('formatted_phone_number','N/A')
        count_of_rating= response_json['result'].get('user_ratings_total','N/A')
        rating = response_json['result'].get('rating','N/A')
        url= response_json['result'].get('url','N/A')
        data = {'website': website,
                            'address': address,
                            'city': city,
                            'state': state,
                            'zip_cod': zip_cod,
                            'phone': phone,
                             'count_of_rating':count_of_rating,
                            'rating': rating,
                            'url': url
                            }
    else:
        print('ERROR!!!!')
        data = {'website': 'N/A',
                            'address': 'N/A',
                            'city': 'N/A',
                            'state': 'N/A',
                            'zip_cod': 'N/A',
                            'phone': 'N/A',
                            'count_of_rating': 'N/A',
                            'rating': 'N/A',
                            'url': 'N/A'
                            }
    write_csv(data)
    print("-----------", data)
    return [data['website'],
        data['address'],
        data['city'],
        data['state'],
        data['zip_cod'],
        data['phone'],
        data['count_of_rating'],
        data['rating'],
        data['url']]

def open_file():
    wb=load_workbook('Web Scrape List - example.xlsx')
    result=[]
    ws=wb.worksheets[0]
    index=1
    for row in ws.iter_rows() :
        cells=[cell.value for cell in row]
        if (cells[0] == None):
            break
        result.append([cell.value for cell in row])
        print(index)
        index +=1
    print(result)
    return result

def write_data_excel(row_need_finish):
    wb = Workbook()
    ws = wb.active
    for row in row_need_finish:
        ws.append(row)
    wb.save("Web Scrape List - fill.xlsx")


def main():
    data_from_excel=open_file()

    for i,row in enumerate(data_from_excel):
        if(i==0):
            data_for_write_excel.append(row)
            continue
            #1 TO 1 ENROLLMENT EDGEWATER MD Insurance
            #1752 FINANCIAL SOLUTIONS RICHMOND VA Insurance
        #https://maps.googleapis.com/maps/api/place/textsearch/json?query=1 TO 1 ENROLLMENT EDGEWATER MD Insurance&key=AIzaSyBLQ3gMYryO8fl_ZBRTpcIiDyNvNT7Maec
        url = 'https://maps.googleapis.com/maps/api/place/textsearch/json?query={}&key=AIzaSyBLQ3gMYryO8fl_ZBRTpcIiDyNvNT7Maec'.format(row[4].strip())
        print('----start')
        print(row[4])
        print(url)
        Unique_ID=row[0]
        Company_Name = row[1]
        Listing_City = row[2]
        Listing_State = row[3]
        Search_Term = row[4]
        data_firts=[Unique_ID,Company_Name,Listing_City,Listing_State,Search_Term]
        response_json = get_page_GM(url)
        time.sleep(1)
        status=response_json['status']
        print(status)
        if (status!='OK'):
            place_id=None
        else:
            place_id=response_json['results'][0]['place_id']
        #print(place_id)
        #ChIJaSzerOQUsYkRVHibpEQMErQ
        #https://maps.googleapis.com/maps/api/place/details/json?place_id=ChIJaSzerOQUsYkRVHibpEQMErQ&language=en&key=AIzaSyBLQ3gMYryO8fl_ZBRTpcIiDyNvNT7Maec
        data_two=get_evry_place(place_id)
        data_for_write_excel.append( data_firts+data_two)

    print(data_for_write_excel)
    write_data_excel(data_for_write_excel)

if __name__ == '__main__':
    main()
